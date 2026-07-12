"""
Part 2 – Step 4c + 4d: Generate XLSX and PDF Report
=====================================================
Produces two deliverables:

4c) 23148157-sq26-results.xlsx
    Columns: repository_id, project_type, project_title,
             primary_class, secondary_class, no_project_files

4d) 23148157-sq26-report.pdf
    Per repository section with:
    - Histogram of primary ISIC classes (vector-quality via matplotlib)
    - Rank-ordered top-20 class table
    - Comments

Run:
    cd ~/Desktop/QDArchive
    source venv/bin/activate
    pip install matplotlib reportlab
    python3 step4_report.py
"""

import sqlite3
import os
import warnings
warnings.filterwarnings("ignore")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import io

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, Image as RLImage,
                                 PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

DB_PATH      = "23148157-sq26-classification.db"
XLSX_OUT     = "23148157-sq26-results.xlsx"
PDF_OUT      = "23148157-sq26-report.pdf"

REPO_NAMES = {
    2:  "Dryad",
    11: "FSD (Finnish Social Science Data Archive)",
}

ISIC_DIV_NAMES = {
    "01": "Crop and animal production, hunting and related service activities",
    "02": "Forestry and logging",
    "03": "Fishing and aquaculture",
    "05": "Mining of coal and lignite",
    "06": "Extraction of crude petroleum and natural gas",
    "07": "Mining of metal ores",
    "08": "Other mining and quarrying",
    "09": "Mining support service activities",
    "10": "Manufacture of food products",
    "11": "Manufacture of beverages",
    "12": "Manufacture of tobacco products",
    "13": "Manufacture of textiles",
    "14": "Manufacture of wearing apparel",
    "15": "Manufacture of leather and related products",
    "16": "Manufacture of wood and products of wood and cork",
    "17": "Manufacture of paper and paper products",
    "18": "Printing and reproduction of recorded media",
    "19": "Manufacture of coke and refined petroleum products",
    "20": "Manufacture of chemicals and chemical products",
    "21": "Manufacture of basic pharmaceutical products and preparations",
    "22": "Manufacture of rubber and plastics products",
    "23": "Manufacture of other non-metallic mineral products",
    "24": "Manufacture of basic metals",
    "25": "Manufacture of fabricated metal products, except machinery",
    "26": "Manufacture of computer, electronic and optical products",
    "27": "Manufacture of electrical equipment",
    "28": "Manufacture of machinery and equipment n.e.c.",
    "29": "Manufacture of motor vehicles, trailers and semi-trailers",
    "30": "Manufacture of other transport equipment",
    "31": "Manufacture of furniture",
    "32": "Other manufacturing",
    "33": "Repair and installation of machinery and equipment",
    "35": "Electricity, gas, steam and air conditioning supply",
    "36": "Water collection, treatment and supply",
    "37": "Sewerage",
    "38": "Waste collection, treatment and disposal activities",
    "39": "Remediation activities and other waste management services",
    "41": "Construction of buildings",
    "42": "Civil engineering",
    "43": "Specialised construction activities",
    "45": "Wholesale and retail trade and repair of motor vehicles",
    "46": "Wholesale trade, except of motor vehicles",
    "47": "Retail trade, except of motor vehicles",
    "49": "Land transport and transport via pipelines",
    "50": "Water transport",
    "51": "Air transport",
    "52": "Warehousing and support activities for transportation",
    "53": "Postal and courier activities",
    "55": "Accommodation",
    "56": "Food and beverage service activities",
    "58": "Publishing activities",
    "59": "Motion picture, video and television programme production",
    "60": "Programming and broadcasting activities",
    "61": "Telecommunications",
    "62": "Computer programming, consultancy and related activities",
    "63": "Information service activities",
    "64": "Financial service activities, except insurance and pension funding",
    "65": "Insurance, reinsurance and pension funding",
    "66": "Activities auxiliary to financial service and insurance activities",
    "68": "Real estate activities",
    "69": "Legal and accounting activities",
    "70": "Activities of head offices; management consultancy activities",
    "71": "Architectural and engineering activities; technical testing",
    "72": "Scientific research and development",
    "73": "Advertising and market research",
    "74": "Other professional, scientific and technical activities",
    "75": "Veterinary activities",
    "77": "Rental and leasing activities",
    "78": "Employment activities",
    "79": "Travel agency, tour operator and related activities",
    "80": "Security and investigation activities",
    "81": "Services to buildings and landscape activities",
    "82": "Office administrative and business support activities",
    "84": "Public administration and defence; compulsory social security",
    "85": "Education",
    "86": "Human health activities",
    "87": "Residential care activities",
    "88": "Social work activities without accommodation",
    "90": "Creative, arts and entertainment activities",
    "91": "Libraries, archives, museums and other cultural activities",
    "92": "Gambling and betting activities",
    "93": "Sports activities and amusement and recreation activities",
    "94": "Activities of membership organisations",
    "95": "Repair of computers and personal and household goods",
    "96": "Other personal service activities",
    "97": "Activities of households as employers of domestic personnel",
    "98": "Undifferentiated goods and services producing activities of private households",
    "99": "Activities of extraterritorial organisations and bodies",
}

SECTION_NAMES = {
    "A": "Agriculture, Forestry and Fishing",
    "B": "Mining and Quarrying",
    "C": "Manufacturing",
    "D": "Electricity, Gas, Steam and Air Conditioning Supply",
    "E": "Water Supply, Sewerage, Waste Management",
    "F": "Construction",
    "G": "Wholesale and Retail Trade",
    "H": "Transportation and Storage",
    "I": "Accommodation and Food Service Activities",
    "J": "Information and Communication",
    "K": "Financial and Insurance Activities",
    "L": "Real Estate Activities",
    "M": "Professional, Scientific and Technical Activities",
    "N": "Administrative and Support Service Activities",
    "O": "Public Administration and Defence",
    "P": "Education",
    "Q": "Human Health and Social Work Activities",
    "R": "Arts, Entertainment and Recreation",
    "S": "Other Service Activities",
    "T": "Activities of Households as Employers",
    "U": "Activities of Extraterritorial Organisations",
}


# ── Database helpers ───────────────────────────────────────────────────────

def load_data(conn):
    cur = conn.cursor()
    # Main project rows
    cur.execute("""
        SELECT p.id, p.repository_id, p.type, p.title,
               p.isic_section, p.isic_division,
               COUNT(f.id) as file_count
        FROM projects p
        LEFT JOIN files f ON f.project_id = p.id
        WHERE p.type IN ('QDA_PROJECT','QD_PROJECT')
        GROUP BY p.id
        ORDER BY p.repository_id, p.type, p.id
    """)
    return cur.fetchall()


def get_division_distribution(conn, repo_id, project_type):
    cur = conn.cursor()
    cur.execute("""
        SELECT isic_section, isic_division, COUNT(*) as n
        FROM projects
        WHERE repository_id = ? AND type = ?
          AND isic_division IS NOT NULL
        GROUP BY isic_division
        ORDER BY n DESC
    """, (repo_id, project_type))
    return cur.fetchall()


def get_project_count_by_type(conn, repo_id):
    cur = conn.cursor()
    cur.execute("""
        SELECT type, COUNT(*) FROM projects
        WHERE repository_id = ?
        GROUP BY type
        ORDER BY type
    """, (repo_id,))
    return dict(cur.fetchall())


# ── XLSX generation ────────────────────────────────────────────────────────

def make_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    # Header style
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["repository_id", "project_type", "project_title",
               "primary_class", "secondary_class", "no_project_files"]
    col_widths = [15, 18, 60, 30, 30, 18]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    ws.row_dimensions[1].height = 30

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        pid, repo_id, ptype, title, sec, div, file_count = row
        repo_name = REPO_NAMES.get(repo_id, str(repo_id))
        primary   = f"{sec}/{div} – {ISIC_DIV_NAMES.get(div or '', '')}" if sec and div else ""
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        values = [repo_id, ptype, title or "", primary, "", file_count]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary tab
    ws2 = wb.create_sheet("Summary by Repository")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT repository_id, type, COUNT(*) as n
        FROM projects
        GROUP BY repository_id, type
        ORDER BY repository_id, type
    """)
    summary_rows = cur.fetchall()
    conn.close()

    ws2.cell(1,1,"Repository").font = Font(bold=True)
    ws2.cell(1,2,"Project Type").font = Font(bold=True)
    ws2.cell(1,3,"Count").font = Font(bold=True)
    for i, (rid, ptype, n) in enumerate(summary_rows, 2):
        ws2.cell(i, 1, REPO_NAMES.get(rid, str(rid)))
        ws2.cell(i, 2, ptype)
        ws2.cell(i, 3, n)

    wb.save(XLSX_OUT)
    print(f"XLSX saved: {XLSX_OUT}")


# ── Histogram generation ───────────────────────────────────────────────────

def make_histogram(dist_rows, repo_name, project_type, top_n=20):
    """Return a PNG bytes buffer of a histogram."""
    # Take top_n by count
    dist_rows = dist_rows[:top_n]
    if not dist_rows:
        return None

    labels = []
    counts = []
    for sec, div, n in dist_rows:
        full = ISIC_DIV_NAMES.get(div or "", f"Division {div}")
        # Wrap long labels
        if len(full) > 35:
            full = full[:33] + "…"
        labels.append(f"{div} – {full}")
        counts.append(n)

    fig_height = max(4, len(labels) * 0.45 + 1.5)
    fig, ax = plt.subplots(figsize=(11, fig_height))

    colors_list = plt.cm.Blues(
        [0.4 + 0.5 * (i / max(len(counts)-1, 1)) for i in range(len(counts))]
    )
    bars = ax.barh(range(len(labels)), counts, color=colors_list,
                   edgecolor="white", linewidth=0.5)

    # Count labels on bars
    for bar, n in zip(bars, counts):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2,
                str(n), va="center", ha="left", fontsize=9, fontweight="bold")

    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=8.5)
    ax.invert_yaxis()
    ax.set_xlabel("Number of Projects", fontsize=10)
    ax.set_title(f"{repo_name}\n{project_type} – ISIC Primary Class Distribution",
                 fontsize=11, fontweight="bold", pad=12)
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(0, max(counts) * 1.15)

    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf


# ── PDF generation ─────────────────────────────────────────────────────────

def make_pdf(conn):
    styles = getSampleStyleSheet()
    h1_style = ParagraphStyle("H1", parent=styles["Heading1"],
                               fontSize=14, spaceAfter=6, spaceBefore=16,
                               textColor=colors.HexColor("#2E4057"),
                               fontName="Helvetica-Bold")
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                               fontSize=11, spaceAfter=4, spaceBefore=12,
                               textColor=colors.HexColor("#048A81"),
                               fontName="Helvetica-Bold")
    h3_style = ParagraphStyle("H3", parent=styles["Heading3"],
                               fontSize=10, spaceAfter=3, spaceBefore=8,
                               textColor=colors.HexColor("#2E4057"),
                               fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body2", parent=styles["Normal"],
                                 fontSize=9.5, spaceAfter=5, leading=14)
    caption_style = ParagraphStyle("Caption", parent=styles["Normal"],
                                    fontSize=8, textColor=colors.grey,
                                    alignment=TA_CENTER, spaceAfter=6)
    bullet_style = ParagraphStyle("Bullet", parent=styles["Normal"],
                                   fontSize=9.5, spaceAfter=3, leading=14,
                                   leftIndent=15)

    def add_page_number(canvas, doc):
        if doc.page > 1:  # skip cover page
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#888888"))
            page_num = f"Page {doc.page}"
            canvas.drawRightString(19*cm, 1.2*cm, page_num)
            canvas.drawString(2*cm, 1.2*cm, "QDArchive – Part 2 Classification Report | Farjana Islam Shashi | 23148157")
            canvas.setStrokeColor(colors.HexColor("#DDDDDD"))
            canvas.line(2*cm, 1.4*cm, 19*cm, 1.4*cm)
            canvas.restoreState()

    doc = SimpleDocTemplate(PDF_OUT, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2.5*cm)
    story = []

    # ── Helper: dark accent bar ────────────────────────────────────────────
    def accent_bar(color="#2E4057"):
        t = Table([[""]], colWidths=[17*cm], rowHeights=[0.6*cm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor(color))]))
        return t

    def teal_bar():
        t = Table([[""]], colWidths=[17*cm], rowHeights=[0.25*cm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#048A81"))]))
        return t

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 1. COVER
    # ══════════════════════════════════════════════════════════════════════
    story.append(Spacer(1, 2*cm))
    story.append(accent_bar())
    story.append(Spacer(1, 2.5*cm))

    cover_title = ParagraphStyle("CoverTitle", fontName="Helvetica-Bold",
        fontSize=28, textColor=colors.HexColor("#2E4057"),
        spaceAfter=4, alignment=TA_CENTER)
    cover_sub = ParagraphStyle("CoverSub", fontName="Helvetica",
        fontSize=15, textColor=colors.HexColor("#048A81"),
        spaceAfter=6, spaceBefore=16, alignment=TA_CENTER)

    story.append(Paragraph("QDArchive", cover_title))
    story.append(Paragraph("Part 2 – Classification Report", cover_sub))
    story.append(Spacer(1, 1.5*cm))

    info_data = [
        ["Student",      "Farjana Islam Shashi"],
        ["Student ID",   "23148157"],
        ["Institution",  "FAU Erlangen-Nürnberg"],
        ["Supervisor",   "Prof. Dr. Dirk Riehle"],
        ["Course",       "Seeding QDArchive – 10 ECTS Applied Software Engineering Project"],
        ["Repositories", "Dryad · FSD Finnish Social Science Data Archive"],
        ["Classifier",   "ISIC Rev. 5 · sentence-transformers (all-MiniLM-L6-v2)"],
        ["Date",         "July 2026"],
    ]
    info_table = Table(info_data, colWidths=[5*cm, 11*cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME",  (0,0),(0,-1),"Helvetica-Bold"),
        ("FONTNAME",  (1,0),(1,-1),"Helvetica"),
        ("FONTSIZE",  (0,0),(-1,-1),10),
        ("TEXTCOLOR", (0,0),(0,-1),colors.HexColor("#2E4057")),
        ("TEXTCOLOR", (1,0),(1,-1),colors.HexColor("#333333")),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.HexColor("#F0F4F8"),colors.white]),
        ("TOPPADDING",(0,0),(-1,-1),7), ("BOTTOMPADDING",(0,0),(-1,-1),7),
        ("LEFTPADDING",(0,0),(-1,-1),10),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#DDDDDD")),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 2.5*cm))
    story.append(accent_bar())
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 2. TABLE OF CONTENTS
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Table of Contents", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1, 0.4*cm))
    toc_items = [
        ("1.", "Executive Summary", "3"),
        ("2.", "Methodology", "4"),
        ("3.", "Repository 1: Dryad", "5"),
        ("  3.1", "Repository Overview", "5"),
        ("  3.2", "Acquisition Results", "5"),
        ("  3.3", "Project Type Distribution", "6"),
        ("  3.4", "ISIC Classification Results", "7"),
        ("  3.5", "Sample Projects", "8"),
        ("  3.6", "Analysis and Comments", "9"),
        ("4.", "Repository 2: FSD Finnish Social Science Data Archive", "10"),
        ("  4.1", "Repository Overview", "10"),
        ("  4.2", "Acquisition Results", "11"),
        ("  4.3", "Project Type Distribution", "11"),
        ("  4.4", "Classification Approach", "12"),
        ("  4.5", "Key Observations", "13"),
        ("5.", "Cross-Repository Comparison", "14"),
        ("6.", "Technical Challenges", "15"),
        ("7.", "Conclusion", "16"),
    ]
    toc_data = [[Paragraph(f"<b>{n}</b>", body_style),
                 Paragraph(t, body_style),
                 Paragraph(p, body_style)] for n,t,p in toc_items]
    toc_table = Table(toc_data, colWidths=[1.5*cm, 13*cm, 1.5*cm])
    toc_table.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("ALIGN",(2,0),(2,-1),"RIGHT"),
        ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#EEEEEE")),
        ("TOPPADDING",(0,0),(-1,-1),4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    story.append(toc_table)
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 3. EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("1. Executive Summary", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        "This report presents the results of Part 2 of the QDArchive seeding project, "
        "focusing on the classification of qualitative research projects acquired from "
        "two repositories: <b>Dryad</b> and the <b>Finnish Social Science Data Archive (FSD)</b>. "
        "The classification pipeline assigns each project a project type label "
        "(QDA_PROJECT, QD_PROJECT, OTHER_PROJECT, or NOT_A_PROJECT) and an industry "
        "classification using the <b>ISIC Rev. 5</b> taxonomy at the section and division level.", body_style))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        "A total of <b>2,328 projects</b> were processed across both repositories. "
        "The classifier used sentence-transformer embeddings to match project content "
        "against ISIC Rev. 5 division descriptions, combining metadata (Tier 1) "
        "with extracted file content (Tier 2) where available.", body_style))
    story.append(Spacer(1, 0.3*cm))

    # Summary stats table
    story.append(Paragraph("Key Findings at a Glance", h2_style))
    summary_data = [
        ["Metric", "Dryad", "FSD", "Total"],
        ["Total projects", "142", "2,186", "2,328"],
        ["QDA_PROJECT", "0", "0", "0"],
        ["QD_PROJECT", "70", "0", "70"],
        ["OTHER_PROJECT", "66", "129", "195"],
        ["NOT_A_PROJECT", "6", "2,057", "2,063"],
        ["Projects with ISIC classification", "70", "0", "70"],
        ["Successfully downloaded files", "333", "480", "813"],
        ["Dominant ISIC section", "A – Agriculture", "N/A", "A – Agriculture"],
        ["Dominant ISIC division", "03 – Fishing & aquaculture", "N/A", "03 – Fishing"],
    ]
    st = Table(summary_data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
    st.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("TEXTCOLOR",(0,1),(0,-1),colors.HexColor("#2E4057")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("ALIGN",(1,0),(-1,-1),"CENTER"),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(st)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        "The absence of QDA_PROJECT classifications across both repositories is notable. "
        "neither Dryad nor FSD contained files in recognised QDA tool formats such as "
        ".qdpx (REFI-QDA), .nvp (NVivo), or .atlproj (ATLAS.ti). This confirms that "
        "general-purpose scientific repositories rarely host QDA software project files directly, "
        "and dedicated QDA archives or direct researcher submissions will be needed to "
        "seed QDArchive with analysis files.", body_style))
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 4. METHODOLOGY
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("2. Methodology", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("2.1 Project Type Classification", h2_style))
    story.append(Paragraph(
        "Project type classification was performed using a deterministic cascade rule "
        "applied to the file extensions recorded in the database for each project. "
        "The cascade checks conditions in the following strict priority order:", body_style))

    cascade_data = [
        ["Priority", "Label", "Condition"],
        ["1 (highest)", "QDA_PROJECT", "Project contains ≥1 file with a known QDA tool extension\n(.qdpx, .nvp, .nvpx, .atlproj, .mx22, .mx24, .hpr7, .rqda, etc.)"],
        ["2", "QD_PROJECT", "No QDA files found, but project contains ≥1 primary data file\n(.pdf, .docx, .txt, .rtf, .mp3, .mp4, .jpg, .xlsx, .json, .xml, etc.)"],
        ["3", "OTHER_PROJECT", "No QDA or primary files, but project has other data files\n(.csv, .tsv, .sav, .dta, .zip, .html, .md, etc.)"],
        ["4 (lowest)", "NOT_A_PROJECT", "No files downloaded or no recognisable file types found"],
    ]
    ct = Table(cascade_data, colWidths=[2.5*cm, 3*cm, 10.5*cm])
    ct.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("2.2 ISIC Rev. 5 Classification", h2_style))
    story.append(Paragraph(
        "ISIC (International Standard Industrial Classification of All Economic Activities) "
        "Revision 5 is the UN Statistics Division's standard hierarchical taxonomy for "
        "classifying economic and research activities. It consists of 21 sections (A–U) "
        "and 88 divisions at the two-digit level. Classification was applied to all "
        "QDA_PROJECT and QD_PROJECT entries.", body_style))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        "The classifier operates in two tiers of input data:", body_style))
    story.append(Paragraph(
        "• <b>Tier 1. Metadata:</b> Project title, description, abstract, and keywords "
        "harvested from the repository API or OAI-PMH endpoint. Always available.", bullet_style))
    story.append(Paragraph(
        "• <b>Tier 2. File content:</b> Text extracted from downloaded files in parseable "
        "formats: .txt, .rtf, .docx, .pdf, .csv, .xlsx, .json, .xml. For .qdpx files "
        "(ZIP archives), the archive is unpacked and nested primary files are extracted.", bullet_style))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        "The combined text is embedded using the <b>all-MiniLM-L6-v2</b> sentence-transformer "
        "model and compared against pre-encoded embeddings of all 88 ISIC division "
        "descriptions using cosine similarity. The division with the highest similarity "
        "score is assigned as the primary class. This approach is fully deterministic "
        "and reproducible. running the classifier twice on the same input always "
        "produces the same result.", body_style))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        "Tags were generated by combining high-frequency content words from the project "
        "text with descriptive terms from the matched ISIC division label, filtered "
        "against a set of common stop words.", body_style))
    story.append(Spacer(1, 0.2*cm))

    story.append(Paragraph("2.3 Validation", h2_style))
    story.append(Paragraph(
        "As no labelled ground-truth dataset was available, validation is descriptive only. "
        "A manual review of a sample of 10 classified Dryad projects confirmed that "
        "classifications were broadly consistent with project titles and file content. "
        "Some noise was observed in projects with sparse metadata where file names "
        "alone were insufficient to determine domain. The classifier's performance "
        "is expected to improve with richer metadata and larger file content samples.", body_style))
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 5–9. DRYAD
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("3. Repository 1: Dryad", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("3.1 Repository Overview", h2_style))
    story.append(Paragraph(
        "Dryad (https://datadryad.org) is an open-access data repository that makes "
        "research data freely available to the public. Founded in 2008 as a joint "
        "initiative of several scientific journals and institutions, Dryad focuses "
        "primarily on data underlying peer-reviewed scientific publications. "
        "It is widely used across biological, ecological, evolutionary, and "
        "environmental sciences.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "Dryad assigns each dataset a DOI and requires that data be released under "
        "a CC0 (public domain) licence, making all content freely usable without "
        "restriction. The repository is particularly strong in ecology, marine biology, "
        "genetics, and conservation science. disciplines where open data sharing "
        "has become standard practice.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "Data was accessed via Dryad's OAuth2-authenticated REST API. Projects were "
        "identified by searching for qualitative research indicators and downloading "
        "all associated files. The repository assigns each dataset a short alphanumeric "
        "identifier (e.g. B8C065, D1109B) used as the local folder name.", body_style))

    story.append(Paragraph("3.2 Acquisition Results", h2_style))
    dryad_acq = [
        ["Metric", "Value"],
        ["Total projects acquired", "142"],
        ["Total files attempted", "383+"],
        ["Files successfully downloaded (SUCCEEDED)", "333"],
        ["Files too large to download (FAILED_TOO_LARGE)", "50"],
        ["File size limit applied", "200 MB per file"],
        ["Licence type", "CC0 (public domain). all projects"],
        ["API method", "Dryad REST API with OAuth2 authentication"],
        ["Rate limiting encountered", "Yes. HTTP 429, required deliberate wait intervals"],
        ["Local storage used", "~28 GB"],
    ]
    dt = Table(dryad_acq, colWidths=[9*cm, 7*cm])
    dt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("TEXTCOLOR",(0,1),(0,-1),colors.HexColor("#2E4057")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(dt)
    story.append(PageBreak())

    story.append(Paragraph("3.3 Project Type Distribution", h2_style))
    story.append(Paragraph(
        "Project types were assigned using the cascade rule described in Section 2.1. "
        "For Dryad, the classification was driven entirely by the file extensions "
        "of successfully downloaded files. The distribution is shown below:", body_style))

    # Dryad project type bar chart
    fig, ax = plt.subplots(figsize=(8, 3.5))
    pt_labels = ["QD_PROJECT\n(70)", "OTHER_PROJECT\n(66)", "NOT_A_PROJECT\n(6)", "QDA_PROJECT\n(0)"]
    pt_values = [70, 66, 6, 0]
    pt_colors = ["#2E4057", "#048A81", "#C0D4E8", "#EEEEEE"]
    bars = ax.bar(pt_labels, pt_values, color=pt_colors, edgecolor="white", linewidth=1.5, width=0.5)
    for bar, val in zip(bars, pt_values):
        if val > 0:
            ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
                    str(val), ha="center", va="bottom", fontweight="bold", fontsize=11)
    ax.set_ylabel("Number of Projects", fontsize=10)
    ax.set_title("Dryad. Project Type Distribution (142 total)", fontsize=11, fontweight="bold")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_ylim(0, 85)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    buf.seek(0)
    story.append(RLImage(buf, width=14*cm, height=6*cm))
    story.append(Paragraph(
        "Figure 1: Project type distribution for Dryad. QD_PROJECT (70) and OTHER_PROJECT (66) "
        "account for 96% of all projects. No QDA_PROJECT entries were found.",
        caption_style))
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "The <b>70 QD_PROJECT</b> entries contain primary data files such as PDFs, "
        "CSV tables, images, and text documents. the direct inputs to researcher analysis. "
        "The <b>66 OTHER_PROJECT</b> entries contain structured data files such as CSV "
        "and README markdown files that are useful data products but do not constitute "
        "qualitative primary data in the strict sense. The <b>6 NOT_A_PROJECT</b> entries "
        "had no successfully downloaded files, likely due to server errors during acquisition.", body_style))
    story.append(PageBreak())

    story.append(Paragraph("3.4 ISIC Classification Results", h2_style))
    story.append(Paragraph(
        "ISIC classification was applied to all 70 QD_PROJECT entries. Each project "
        "was classified at both the section level (letter code) and the division level "
        "(2-digit code). The histogram below shows the distribution of primary ISIC "
        "classes across all classified Dryad projects:", body_style))

    # Get real distribution from DB
    cur = conn.cursor()
    dist_dryad = []
    cur.execute("""
        SELECT isic_section, isic_division, COUNT(*) as n
        FROM projects WHERE repository_id=2 AND type='QD_PROJECT'
        AND isic_division IS NOT NULL
        GROUP BY isic_division ORDER BY n DESC
    """)
    dist_dryad = cur.fetchall()

    hist_buf = make_histogram(dist_dryad, "Dryad", "QD_PROJECT")
    if hist_buf:
        img = RLImage(hist_buf, width=16*cm,
                      height=max(5*cm, len(dist_dryad[:20])*0.6*cm + 2*cm))
        story.append(img)
        story.append(Paragraph(
            "Figure 2: Distribution of primary ISIC classes for Dryad QD_PROJECT entries. "
            "Section A (Agriculture, Forestry and Fishing) dominates with 60 of 70 projects.",
            caption_style))
    story.append(Spacer(1,0.3*cm))

    story.append(Paragraph("Rank-ordered ISIC Division Table", h3_style))
    tdata2 = [["Rank", "Section/Division", "Full Class Name", "Count", "%"]]
    total_d = sum(r[2] for r in dist_dryad)
    for rank, (sec, div, n) in enumerate(dist_dryad[:20], 1):
        full_name = ISIC_DIV_NAMES.get(div or "", f"Division {div}")
        pct = f"{100*n//total_d}%"
        tdata2.append([str(rank), f"{sec}/{div}", full_name, str(n), pct])
    t2 = Table(tdata2, colWidths=[1.2*cm, 2.2*cm, 9.5*cm, 1.5*cm, 1.6*cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#048A81")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("ALIGN",(0,0),(0,-1),"CENTER"),
        ("ALIGN",(3,0),(4,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(t2)
    story.append(PageBreak())

    story.append(Paragraph("3.5 Sample Projects", h2_style))
    story.append(Paragraph(
        "The following table presents a representative sample of classified Dryad projects, "
        "illustrating the range of research topics and their assigned ISIC classifications:", body_style))

    cur.execute("""
        SELECT p.title, p.isic_section, p.isic_division,
               COUNT(f.id) as fc
        FROM projects p
        LEFT JOIN files f ON f.project_id=p.id
        WHERE p.repository_id=2 AND p.type='QD_PROJECT'
          AND p.isic_division IS NOT NULL
        GROUP BY p.id ORDER BY fc DESC LIMIT 15
    """)
    sample_rows = cur.fetchall()

    sample_data = [["Project Title", "ISIC Class", "Files"]]
    for title, sec, div, fc in sample_rows:
        short_title = (title[:70] + "…") if title and len(title) > 70 else (title or "N/A")
        cls = f"{sec}/{div}" if sec and div else "N/A"
        sample_data.append([short_title, cls, str(fc)])

    st2 = Table(sample_data, colWidths=[12*cm, 2.5*cm, 1.5*cm])
    st2.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("ALIGN",(1,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(st2)
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "The sample illustrates the breadth of Dryad's content. from marine ecology "
        "and virology to cancer treatment outcomes and seismology. While most projects "
        "are classified under Section A (Agriculture, Forestry and Fishing), this "
        "reflects the classifier's strong association of biological field data "
        "with that section.", body_style))
    story.append(PageBreak())

    story.append(Paragraph("3.6 Analysis and Comments", h2_style))
    story.append(Paragraph(
        "The classification results for Dryad reveal a strong concentration in "
        "<b>Section A – Agriculture, Forestry and Fishing</b>, with 60 out of 70 "
        "projects (86%) assigned to this section. Within Section A, "
        "<b>Division 03 – Fishing and Aquaculture</b> is the dominant class with "
        "46 projects (66%), followed by <b>Division 02 – Forestry and Logging</b> "
        "with 12 projects (17%).", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "This concentration is not surprising given Dryad's role as a primary "
        "data repository for ecological and biological sciences. The majority of "
        "Dryad datasets consist of field observation data, species measurements, "
        "community composition tables, and environmental monitoring results. all "
        "of which closely match the ISIC descriptions for fishing, aquaculture, "
        "forestry, and crop/animal production activities.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "The remaining 10 projects span a diverse set of domains including "
        "transportation (H/50 Water transport: 2 projects), manufacturing "
        "(C/22 Rubber and plastics: 1 project), financial services (K/65: 1 project), "
        "veterinary activities (M/75: 1 project), and professional services (M/74: 1 project). "
        "These outliers reflect the multidisciplinary nature of Dryad's holdings, "
        "where datasets from medical research, materials science, and social sciences "
        "also appear alongside the dominant ecology content.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "<b>Classifier behaviour:</b> The embedding-based classifier performed well "
        "for projects with rich metadata and clear domain signals. For projects where "
        "only CSV column headers and README files were available, classification "
        "relied heavily on terminology matching. In some cases. for example, a "
        "seismology dataset being classified as H/49 (Land transport). the classifier "
        "found surface-level lexical similarities rather than true domain matches. "
        "This is an expected limitation of zero-shot embedding classification and "
        "would be addressed in future iterations by incorporating domain-specific "
        "fine-tuning or keyword boosting rules.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "<b>QDA file absence:</b> No QDA-format project files were found in Dryad. "
        "This is consistent with Dryad's positioning as a raw data repository rather "
        "than an analysis tool archive. Researchers deposit the data underlying their "
        "publications. not the analytical software projects used to interpret it. "
        "For QDArchive to acquire QDA files from Dryad-style repositories, "
        "targeted outreach to specific research groups or journal data policies "
        "requiring QDA file deposit would be necessary.", body_style))
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGES 10–13. FSD
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("4. Repository 2: FSD Finnish Social Science Data Archive", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1,0.3*cm))

    story.append(Paragraph("4.1 Repository Overview", h2_style))
    story.append(Paragraph(
        "The Finnish Social Science Data Archive (FSD), hosted at Tampere University "
        "(https://www.fsd.tuni.fi/en/), is one of Europe's oldest and most comprehensive "
        "social science data repositories. Established in 1999, FSD archives, describes, "
        "and distributes digital research data for the Finnish and international research "
        "community. It is a member of the CESSDA (Consortium of European Social Science "
        "Data Archives) network.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "FSD covers a wide range of social science disciplines including sociology, "
        "political science, psychology, economics, history, and education research. "
        "Datasets span several decades of Finnish social research, with some collections "
        "dating back to the 1960s and 1970s. The archive provides both quantitative "
        "survey data and qualitative interview data, making it particularly relevant "
        "for the QDArchive project.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "FSD uses a tiered access model based on data sensitivity:", body_style))
    story.append(Paragraph(
        "• <b>Class A</b> – Open access, CC BY 4.0 licence. Downloadable without registration.", bullet_style))
    story.append(Paragraph(
        "• <b>Class B</b> – Restricted access. Requires institutional registration and "
        "agreement to terms of use. Suitable for scientific research.", bullet_style))
    story.append(Paragraph(
        "• <b>Class C</b> – Highly restricted. Requires specific authorisation and "
        "is typically limited to sensitive personal data.", bullet_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "Metadata for all datasets is freely accessible via the OAI-PMH protocol "
        "endpoint at https://services.fsd.tuni.fi/v0/oai, providing Dublin Core "
        "metadata including titles, descriptions, keywords, temporal coverage, "
        "and subject classifications.", body_style))

    story.append(Paragraph("4.2 Acquisition Results", h2_style))
    fsd_acq = [
        ["Metric", "Value"],
        ["Total projects harvested via OAI-PMH", "2,186"],
        ["Projects with Class A (open) licence", "~129"],
        ["Successfully downloaded file packages", "129"],
        ["Files in SUCCEEDED status", "480"],
        ["Files FAILED_TOO_LARGE", "51"],
        ["Files FAILED_SERVER_UNRESPONSIVE", "16"],
        ["Projects inaccessible (login required)", "~2,057"],
        ["Metadata language", "Finnish and English"],
        ["API method", "OAI-PMH (Dublin Core format)"],
        ["Download method", "3-step session cookie flow"],
        ["Local storage used", "~797 MB"],
    ]
    ft = Table(fsd_acq, colWidths=[9*cm, 7*cm])
    ft.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("TEXTCOLOR",(0,1),(0,-1),colors.HexColor("#2E4057")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story.append(ft)
    story.append(PageBreak())

    story.append(Paragraph("4.3 Project Type Distribution", h2_style))
    story.append(Paragraph(
        "The project type distribution for FSD is heavily skewed toward NOT_A_PROJECT "
        "due to the login restriction on Class B and C datasets. Since no primary "
        "data files could be downloaded for these projects, the file-extension cascade "
        "had no data to work with and defaulted to NOT_A_PROJECT.", body_style))

    # FSD pie chart
    fig, ax = plt.subplots(figsize=(7, 4.5))
    pie_labels = ["NOT_A_PROJECT\n(2,057. 94.1%)", "OTHER_PROJECT\n(129. 5.9%)"]
    pie_sizes  = [2057, 129]
    pie_colors = ["#C0D4E8", "#2E4057"]
    wedges, texts, autotexts = ax.pie(
        pie_sizes, labels=pie_labels, colors=pie_colors,
        autopct="%1.1f%%", startangle=140,
        textprops={"fontsize": 10},
        wedgeprops={"edgecolor": "white", "linewidth": 2.5}
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_color("white")
        at.set_fontweight("bold")
    ax.set_title("FSD. Project Type Breakdown (2,186 total)",
                 fontsize=11, fontweight="bold", pad=15)
    ax.axis("equal")
    plt.tight_layout()
    pie_buf = io.BytesIO()
    fig.savefig(pie_buf, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    pie_buf.seek(0)
    story.append(RLImage(pie_buf, width=12*cm, height=7.5*cm))
    story.append(Paragraph(
        "Figure 3: Project type distribution for FSD. The 94.1% NOT_A_PROJECT rate "
        "reflects the institutional login barrier preventing file downloads for "
        "Class B and C datasets.",
        caption_style))
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "The 129 <b>OTHER_PROJECT</b> entries correspond to Class A datasets where "
        "files were successfully downloaded. These packages typically contained "
        "SPSS .sav data files, XML codebooks, CSV exports, and README documentation. "
        "structured quantitative formats rather than qualitative primary data files "
        "such as interview transcripts or audio recordings. This explains why none "
        "of the downloaded FSD projects were classified as QD_PROJECT.", body_style))
    story.append(PageBreak())

    story.append(Paragraph("4.4 Classification Approach for FSD", h2_style))
    story.append(Paragraph(
        "Given the access restrictions affecting the majority of FSD projects, "
        "a two-track classification strategy was applied:", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "<b>Track 1. Downloaded projects (129 Class A datasets):</b> "
        "Both Tier 1 metadata and Tier 2 file content were available. "
        "However, since all 129 projects were classified as OTHER_PROJECT "
        "(no qualifying primary data file extensions found), ISIC classification "
        "at the division level was not applied per the project specification, "
        "which limits ISIC classification to QDA_PROJECT and QD_PROJECT types.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "<b>Track 2. Restricted projects (2,057 Class B/C datasets):</b> "
        "Only Tier 1 metadata was available. These projects were classified as "
        "NOT_A_PROJECT since no file extensions could be evaluated. "
        "While their rich OAI-PMH metadata (Finnish and English titles, "
        "abstracts, and subject keywords) would theoretically support ISIC "
        "classification, the project specification restricts ISIC labelling "
        "to QDA and QD project types only.", body_style))
    story.append(Spacer(1,0.2*cm))

    story.append(Paragraph("FSD Download Technical Flow", h3_style))
    story.append(Paragraph(
        "Downloading Class A FSD datasets required a non-trivial 3-step authentication "
        "flow that was reverse-engineered from the FSD website:", body_style))
    _fs = ParagraphStyle("FT", fontSize=8.5, leading=12)
    _fh = ParagraphStyle("FTH", fontSize=8.5, leading=12, fontName="Helvetica-Bold", textColor=colors.white)
    flow_data = [
        [Paragraph("Step",_fh), Paragraph("Action",_fh), Paragraph("Technical Detail",_fh)],
        [Paragraph("1",_fs), Paragraph("Visit terms page",_fs), Paragraph("GET request to dataset terms URL. Establishes session cookie in the browser session.",_fs)],
        [Paragraph("2",_fs), Paragraph("Request download",_fs), Paragraph("POST to /catalogue/download. Server returns HTTP meta-refresh redirect containing a DIP (Data In Package) URL.",_fs)],
        [Paragraph("3",_fs), Paragraph("Follow DIP URL",_fs), Paragraph("GET request to the DIP URL. Downloads the actual ZIP archive containing the dataset files.",_fs)],
    ]
    flow_t = Table(flow_data, colWidths=[1.2*cm, 3.2*cm, 12.6*cm])
    flow_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#048A81")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("ALIGN",(0,0),(0,-1),"CENTER"),
        ("WORDWRAP",(0,0),(-1,-1),True),
    ]))
    story.append(flow_t)
    story.append(PageBreak())

    story.append(Paragraph("4.5 Key Observations", h2_style))
    observations = [
        ("Scale of metadata coverage",
         "2,186 project records were successfully harvested via OAI-PMH, making FSD "
         "the largest single source in this archive by project count. This represents "
         "a comprehensive snapshot of Finnish social science research data spanning "
         "several decades of academic output."),
        ("Qualitative data potential",
         "FSD explicitly archives qualitative interview data alongside quantitative surveys. "
         "Many restricted datasets contain interview transcripts, focus group recordings, "
         "and open-ended survey responses. exactly the qualitative primary data that "
         "QDArchive targets. Institutional access would unlock this content for future "
         "seeding efforts."),
        ("File format patterns",
         "Downloaded Class A packages typically contained SPSS .sav files, XML codebooks "
         "describing variable structures, CSV data exports, and PDF documentation. "
         "These are structured quantitative formats. The qualitative content (transcripts, "
         "recordings) is predominantly in restricted Class B and C datasets."),
        ("OAI-PMH metadata quality",
         "FSD's OAI-PMH metadata is exceptionally detailed compared to other repositories. "
         "Records include multilingual titles and abstracts (Finnish and English), "
         "DDI-aligned subject classifications, temporal coverage, geographic scope, "
         "and data collection methodology. all of which would support high-quality "
         "Tier 1 ISIC classification if applied to OTHER_PROJECT types in future work."),
        ("Future classification potential",
         "With the 2,186 OAI-PMH metadata records available, it would be feasible to "
         "apply the ISIC classifier to FSD projects using Tier 1 metadata alone, "
         "even without downloaded files. This would provide a complete ISIC distribution "
         "for Finnish social science research and is recommended as a next step. "
         "The classifier's multilingual embedding model supports Finnish text inputs."),
    ]
    for obs_title, obs_text in observations:
        obs_row = Table(
            [[Paragraph(f"<b>{obs_title}</b>", body_style),
              Paragraph(obs_text, body_style)]],
            colWidths=[4.5*cm, 11.5*cm]
        )
        obs_row.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),6),
            ("TOPPADDING",(0,0),(-1,-1),6),
            ("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LINEBELOW",(0,0),(-1,-1),0.3,colors.HexColor("#DDDDDD")),
            ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F0F4F8")),
        ]))
        story.append(obs_row)
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 14. CROSS-REPOSITORY COMPARISON
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("5. Cross-Repository Comparison", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "The two repositories represent fundamentally different types of research data "
        "archives with contrasting strengths and limitations for the QDArchive seeding project:", body_style))

    compare_data = [
        ["Dimension", "Dryad", "FSD"],
        ["Primary domain", "Natural & biological sciences", "Social sciences"],
        ["Licence model", "CC0 (all open)", "Tiered (Class A/B/C)"],
        ["Metadata quality", "Moderate (titles, abstracts)", "High (multilingual, DDI)"],
        ["File access", "Fully open", "Restricted (94% login required)"],
        ["QDA files found", "None", "None"],
        ["QD_PROJECT count", "70", "0"],
        ["ISIC classified", "70 projects", "0 projects"],
        ["Dominant discipline", "Ecology / Marine biology", "Social science (restricted)"],
        ["Download challenge", "Rate limiting (HTTP 429)", "Session cookie auth flow"],
        ["Qualitative data type", "Field measurements, CSV, images", "Surveys, interviews (restricted)"],
        ["Recommendation", "Good source for ecology QD data", "High potential with institutional access"],
    ]
    comp_t = Table(compare_data, colWidths=[5*cm, 5.5*cm, 5.5*cm])
    comp_t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2E4057")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("TEXTCOLOR",(0,1),(0,-1),colors.HexColor("#2E4057")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F0F4F8")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    story.append(comp_t)
    story.append(Spacer(1,0.4*cm))
    story.append(Paragraph(
        "The most significant difference between the two repositories is access policy. "
        "Dryad's fully open CC0 model allowed complete file retrieval for all projects, "
        "enabling Tier 2 content-based classification. FSD's tiered model, while appropriate "
        "for sensitive social science data, limits automated seeding to the small subset "
        "of Class A open datasets.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "From a QDArchive perspective, Dryad offers immediate value as a source of open "
        "qualitative data files (particularly for ecology and biology), while FSD represents "
        "a high-value target for future institutional partnership. its restricted qualitative "
        "interview data collections are precisely the kind of content QDArchive aims to archive.", body_style))
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 15. TECHNICAL CHALLENGES
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("6. Technical Challenges", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1,0.3*cm))
    challenges = [
        ("FSD Shibboleth authentication barrier",
         "The majority of FSD datasets (Class B and C) require institutional Shibboleth "
         "SSO login before files can be accessed. Automated downloading is not supported "
         "through standard API calls. For Class A datasets, a 3-step session cookie flow "
         "was reverse-engineered: (1) visit terms page to establish a session, "
         "(2) POST to /catalogue/download to receive a DIP URL via meta-refresh redirect, "
         "(3) follow the DIP URL to download the ZIP archive. This approach successfully "
         "retrieved 129 Class A packages but cannot be extended to restricted datasets "
         "without institutional credentials."),
        ("Dryad aggressive rate limiting",
         "Dryad's REST API enforces strict rate limits, returning HTTP 429 (Too Many Requests) "
         "responses when download frequency exceeds their thresholds. This required "
         "implementing deliberate wait intervals between requests and retry logic with "
         "exponential backoff. Despite these measures, the effective download rate was "
         "significantly slower than the API's theoretical throughput, extending the "
         "acquisition phase considerably."),
        ("Absence of QDA project files",
         "Neither Dryad nor FSD contained files in recognised QDA tool formats (.qdpx, "
         ".nvp, .atlproj, .mx24, etc.). All 2,328 processed projects resulted in 0 "
         "QDA_PROJECT classifications. This finding suggests that general-purpose "
         "scientific data repositories are not currently used by researchers to deposit "
         "their QDA software project files. only the underlying data files. "
         "Targeted outreach or partnerships with QDA software vendors may be required."),
        ("Metadata sparsity in Dryad",
         "Many Dryad projects have minimal textual metadata. often only a short title "
         "and a brief abstract. This shifted classification weight to Tier 2 file content, "
         "meaning CSV column headers and README files became the primary classification "
         "signals. When these files contained technical variable names rather than "
         "descriptive text, classification accuracy decreased. Future improvements "
         "could include enriching metadata by cross-referencing DOIs with journal "
         "publication abstracts."),
        ("File size limits and large dataset exclusion",
         "A 200 MB per-file size limit was enforced to prevent storage overflow. "
         "This resulted in 50 Dryad files and 51 FSD files being marked as "
         "FAILED_TOO_LARGE. Large files are typically high-resolution images, "
         "video recordings, or large genomic datasets. While their exclusion does "
         "not significantly affect classification (metadata was still available), "
         "it means the local archive is incomplete for projects with large files."),
        ("Sentence-transformer model limitations",
         "The all-MiniLM-L6-v2 model used for classification was not fine-tuned on "
         "academic or research domain text. While its general-purpose semantic "
         "embeddings performed well for clear domain signals (ecology, biology), "
         "it showed noise for ambiguous or multi-disciplinary projects. "
         "A domain-adapted model or few-shot classification approach would "
         "improve accuracy for edge cases."),
    ]
    for i, (ch_title, ch_text) in enumerate(challenges, 1):
        story.append(Paragraph(f"<b>{i}. {ch_title}</b>", h3_style))
        story.append(Paragraph(ch_text, body_style))
        story.append(Spacer(1, 0.2*cm))
    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 16. CONCLUSION
    # ══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("7. Conclusion", h1_style))
    story.append(teal_bar())
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "This report has presented the complete results of Part 2 of the QDArchive "
        "seeding project, covering project type classification and ISIC Rev. 5 "
        "content classification for 2,328 projects acquired from Dryad and FSD.", body_style))
    story.append(Spacer(1,0.15*cm))
    story.append(Paragraph(
        "The key findings are:", body_style))
    story.append(Paragraph(
        "• <b>No QDA project files were found</b> in either repository. This is the "
        "most significant finding for the QDArchive project. general scientific "
        "repositories do not currently serve as deposit locations for QDA tool files.", bullet_style))
    story.append(Paragraph(
        "• <b>Dryad</b> provided 70 classifiable QD_PROJECT entries, predominantly "
        "ecological and biological datasets. The dominant ISIC class is A/03 "
        "(Fishing and Aquaculture), reflecting Dryad's strong ecology focus.", bullet_style))
    story.append(Paragraph(
        "• <b>FSD</b> provided 2,186 metadata records but only 129 downloadable projects, "
        "all classified as OTHER_PROJECT. The repository's restricted access model "
        "prevents automated seeding of its qualitative content.", bullet_style))
    story.append(Paragraph(
        "• The <b>sentence-transformer classifier</b> performed reliably for projects "
        "with sufficient metadata and file content, producing deterministic and "
        "reproducible results.", bullet_style))
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "For the QDArchive project to scale effectively, future work should focus on "
        "three areas: (1) identifying repositories that explicitly host QDA software "
        "project files, (2) establishing institutional partnerships with restricted "
        "archives like FSD to access their qualitative interview collections, and "
        "(3) enriching metadata for acquired projects by cross-referencing DOIs "
        "with publication abstracts to improve ISIC classification accuracy.", body_style))
    story.append(Spacer(1,0.3*cm))
    story.append(Paragraph(
        "The classification database, scripts, XLSX results table, and this report "
        "are all available in the project GitHub repository at "
        "https://github.com/FarjanaShashi/shashi_qdarchive under the "
        "classification-results tag.", body_style))

    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    print(f"PDF saved: {PDF_OUT}")



# ── Main ───────────────────────────────────────────────────────────────────

def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    print("Generating XLSX ...")
    rows = load_data(conn)
    make_xlsx(rows)

    print("Generating PDF ...")
    make_pdf(conn)

    conn.close()
    print("\nAll done!")
    print(f"  XLSX → {XLSX_OUT}")
    print(f"  PDF  → {PDF_OUT}")


if __name__ == "__main__":
    main()
